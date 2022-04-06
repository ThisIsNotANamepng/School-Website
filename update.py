#!/usr/bin/env python
#This takes the xlsx file from /uploads, edits the /Data pages to reflect upadted information from xlsx file. 
#pip3 install beautifulsoup4
#pip3 install lxml
import os
import openpyxl
import glob
import json
from bs4 import BeautifulSoup
import lxml

error = []#this will collect any errors that happens
filenames = []
teachernames = []

os.chdir('../')

for file in glob.glob("uploads/*.xlsx"):
  path = (file)#Needs to open the one file in 'uploads' folder
  delete = path


wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
ws = sheet_obj
 
sheet_object = wb_obj.active
max_rows = sheet_object.max_row#finding how many rows there are

os.chdir('Data')
i = 2#if i=1, it would use "21/22" (the year) as well
while i <= max_rows:

  rooms = []
  subjects = []
  classes = []
  #The three empties every time it runs, clearing it for the next name

  cell_obj = sheet_obj.cell(row = i, column = 1)
  if (cell_obj.value is None) or (cell_obj.value == "GYM") or (cell_obj.value == "DANCE"):#filters out empty, "DANCE", and "GYM" results
    a = 1
  else:
    txt = cell_obj.value
    x = "RM " in txt
    
    if (x == True):#filters out room number results


      
      a = 1 
    else:
        #This is where we get the room number, it tales the cell below it and gets rif of "RM ", and then appends it to 'rooms'-----
        room_number = sheet_obj.cell(row = i+1, column = 1)
        room_number = room_number.value
        rooms.append(room_number.replace("RM ", ""))
        #------
        #This is where we get the classes-------
        ii = 1
        while ii <=10:#10 because that's how many columns there are total
          class_test = sheet_obj.cell(row = i, column = ii)
          class_test = class_test.value
          #Now 'class_test' is the value of a cell (i, ii) being the coordinates. We can check to see if it's in 'class' already. If it is, we can ignore it. If it's not we know it's a new class and will add it. But first, we need to check if it's value is "None" and get rid of the "\n(teacher)"
          if class_test != None:
            if "\n" in class_test:
              #delete the "\n" and the parantheses, including everything inside the parantheses
              v = 0
              for v in range(0, len(class_test)):
                if class_test[v] == "(":
                  start = v
                if class_test[v] == ")":
                  end = v
              vv = start
              class_test_list = []
              class_test_list[:0]=class_test
              while vv<=end:
                del class_test_list[start]
                vv = vv+1

              class_test_list.remove("\n")
              
              class_test = "" 
    
              for ele in class_test_list: 
                class_test += ele 
            a = 1
                    
            
            if class_test not in classes and class_test != None:
              class_test = str(class_test)
              classes.append(class_test)
          ii = ii+1

        #deleting the teachers' name from the forst index, somehow that got in there
        classes.pop(0)
        if "PREP" in classes:
          classes.remove("PREP")
      
      
      
      



        #-----------------------------------
           
        searchThis = cell_obj.value
        teacherName = searchThis
        searchThis = searchThis.replace(" ", "")

        teachernames.append(teacherName)
        filenames.append(searchThis)

        #This is where we open the html file and edit it
        os.chdir("..")
        with open('TeacherTemplate', 'r') as f:
      
          contents = f.read()
          
      
          soup = BeautifulSoup(contents, 'lxml')

          
        os.chdir('Data')

        #---------Subejcts
        tag = (soup.find(id='subjects'))#Finds h3 with id: subjects
        #These are lists of all of the classes. If a class is in one of the lists, the subject is the one corresponding to that list
        #make testing class lowercase, classes in subject lists are all lowercase
      
        #If a class comes in that is unrecognized or spelled differently or is a little different from the official name, just add it into the list so in the future when it comes back agin it can be recognized
        
        arts = ["acting 1", "putting on a show", "film acting", "music theatre", "theatre arts", "ib theatre", "advanced acting 2", "advanced acting 3", "advanced acting 4", "school of rock", "choraliers", "concert choir", "treble choir", "ib music sl", "concert band", "symphonic band", "wind symphony", "orchestra", "drawing 1", "painting 1", "sculpture 1", "ceramics 1", "digital photography 1", "digital art", "advanced digital art", "industrial design", "advertising and design", "ap art and design", "advanced sculpture and ceramics 1", "advanced sculpture and ceramics 2", "advanced sculpture and ceramics 3", "advanced photography 2", "advanced photography 3", "design challenge", "advanced drawing and painting 2", "advanced drawing and painting 3"]
        english = ["english 9", "english 10", "acc english 10", "acc english 9", "english 11", "ap english language and composition", "ib english hl", "ib english 2 hl", "english 12", "ap english literature and composition", "intercultural communication", "contemporary nonfiction", "creative writing", "literature and film theory", "seminar digital communications", "ib extended essay", "digital productions"]
        math = ["algebra 1", "geometry", "acc geometry", "trades math", "algebra 2", "acc algebra 2", "college algebra", "precalculus", "ib math sl", "introduction to statistics", "ap statistics", "ap statistics ab", "ap statistics bc", "pie calculus 3", "mathematics in global issues"]
        business = ["career internship", "career portfolio", "education internship", "finance youth apprenticeship", "hospitality youth apprenticeship", "hsb business economics", "hsb business strategies", "hsb principles of business", "hsb principles of finance", "hsb principles managment", "hsb principles marketing", "ib personal and professional skills 1", "ib personal and professional skills 2", "intro to accounting", "marketing youth apprenticeship", "personal finance", "tc social media marketing", "sports and entertainment marketing", "tc college accounting", "tc hospitality", "tc microsoft word and powerpoint", "tc microsoft excel and quickbooks", "tc web design", "yearbook"]
        computerscience = ["elements of game design", "computer science discoveries", "ap computer science principles", "ap computer science", "ap computer science a", "tc programming for the web", "tc mobile app development", "tech internship"]
        language = ["chinese 1", "chinese 2", "chinese 3", "chinese 4", "chinese 5", "ib chinese sl 1", "ib chinese sl 2", "french 1", "french 2", "french 3", "french 4", "ib french sl", "german 1", "german 2", "german 3", "german 4", "german 5", "ib german sl 1", "ib german sl 2", "ib german hl", "spanish 1", "spanish 2", "spanish 3", "spanish 4", "spanish 5", "ib spanish sl 1", "ib spanish hl 1", "ib spanish sl 2", "ib spanish hl 2", "global sustainability"]
        physicalfitness = ["health", "broadfield", "team and individual", "wellness watch", "lifetime pursuits", "dance", "weight training", "advanced fitness", "officiating", "intrduction to leadership"]
        science = ["biology", "chemistry", "acc chemistry", "physics", "earth and space science", "ap chemistry", "ap chem", "ap physics 1", "ap physics c", "ap environmental science", "ib biology hl 1", "ib biology hl 2", "ib chemistry hl 1", "ib chemistry hl 2", "ib physics sl", "ib physics", "pltw human body", "pltw human body systems", "pltw human body interventions", "pltw biological innovations", "global sustainability"]
        teched = ["building trades 1", "building trades 2", "building trades 3", "digital productions", "oconmanufacturing", "oconmanufacturing 1", "oconmanufacturing 2", "advanced welding", "intro to engineering", "intro to engineering design", "introduction to engineering design", "principles of engineering", "civil engineering and architecture", "digital electronics", "computer integrated manufacturing", "engineering design and development", "oconfablab", "industrial design", "consumer auto", "auto 1", "auto", "auto 2", "auto 3"]
        socialstudies = ["humanities 9", "modern world history", "ap human geography", "united states history", "sociology", "ap us history", "ap world history", "ap workd", "ap workd history: modern", "citizenship", "ap american government and politics", "the law", "ap psychology", "aconomics", "psychology", "ib economics sl 1", "ib history of the americas hl 1", "ib history of the americas hl 2", "ib history of the americas", "ib theory of knowledge", "ib theory of knowledge hl 1", "ib theory of knowledge hl 2", "ap seminar", "cin", "ap seminar/cin"]
        for qqq in classes:
          qqq = qqq.lower()
          
          if "student services" in qqq and "Student Services" not in subjects:
            subjects.append("Student Services")
          if qqq in socialstudies and "Social Studies" not in subjects:
            subjects.append("Social Studies")
          if qqq in science and "Science" not in subjects:
            subjects.append("Science")
          if qqq in teched and "Technology Education" not in subjects:
            subjects.append("Technology Education")
          if qqq in physicalfitness and "Physical Fitness" not in subjects:
            subjects.append("Physical Fitness")
          if qqq in language and "World Language" not in subjects:
            subjects.append("World Language")
          if qqq in computerscience and "Computer Science" not in subjects:
            subjects.apppend("Computer Science")
          if qqq in business and "Business" not in subjects:
            subjects.append("Business")
          if qqq in math and "Math" not in subjects:
            subjects.append("Math")
          if qqq in english and "English" not in subjects:
            subjects.append("English")
          if qqq in arts and "Arts" not in subjects:
            subjects.append("Arts")
          if qqq not in arts and qqq not in socialstudies and qqq not in science and qqq not in teched and qqq not in physicalfitness and qqq not in language and qqq not in computerscience and qqq not in business and qqq not in math and qqq not in english and "student services" not in qqq:
            print(error)
            terror = ("Unknown class:  "+qqq)
            error.append(terror)

        if subjects:
          build_subjects = subjects[0]#Starting the loop off with first item
          r = 1
          while r<len(subjects):#couldn't get a for loop to work
            build_subjects = build_subjects+", "+subjects[r]#Concated each item to variable passed form last loop through
            r = r+1
      
          subject_string = '<h3 id="subject">Subjects: '+build_subjects+'</h3>'
          tag.replace_with(subject_string)
        else:
          tag.replace_with('<h3 id="subject"></h3>')
        #---------Classes
    
        tag = (soup.find(id='classes'))
        build_classes = classes[0]#only adding one class---------------------------------------------------------------------------------------------
        c = 1
        while c<len(classes):
          build_classes = build_classes+", "+classes[c]
          c = i+1
    
        class_string = '<h3 id="classes">Classes: '+build_classes+'</h3>'
        tag.replace_with(class_string)
        #---------Rooms
    
        tag = (soup.find(id='rooms'))
        #rooms = ["201", "302"]#Keep room numbers in string, integers would prpbably require some tweaking of the code
        build_rooms = rooms[0]
        g = 1
        while i<len(rooms):
          build_rooms = build_rooms+", "+rooms[g]
          g = i+1
    
        room_string = '<h3 id="rooms">Rooms: '+build_rooms+'</h3>'
        tag.replace_with(room_string)
      
      
        with open(searchThis, "w") as f_output:
    
          f_output.write(soup.prettify(formatter=None)) 
        

  i = i+1
  
if error:
  print("Update completed with the following errors:")
  print(error)
else:
  print("Update was successful without any errors")
os.chdir('../')
os.remove(delete)     #deletes uploaded file