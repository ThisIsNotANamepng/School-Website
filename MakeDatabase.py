#!/usr/bin/env python
#This takes the xlsx file from /uploads, creates the xml file, 
from xml.dom import minidom
import os
import openpyxl
import xml.etree.ElementTree as gfg
import glob
from xml.etree import ElementTree
import json


added = []
removed = []
path = "Data"
dir_list = os.listdir(path)

root = gfg.Element("pages")#xml things


for file in glob.glob("uploads/*.xlsx"):
  path = (file)#Needs to open the one file in 'uploads' folder
  delete = path


wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
ws = sheet_obj
 
sheet_object = wb_obj.active
max_rows = sheet_object.max_row#finding how many rows there are

os.chdir('Data')
i = 2#if i=1, it would use "21/22" (the year) as well
while i <= max_rows:
  cell_obj = sheet_obj.cell(row = i, column = 1)
  if (cell_obj.value is None) or (cell_obj.value == "GYM") or (cell_obj.value == "DANCE"):#filters out empty, "dance", and "gym" results
    a = 1
  else:
    txt = cell_obj.value
    x = "RM " in txt
    if (x == True):#filters out room number results
      a = 1
    else:
      
        m1 = gfg.Element("school")
        root.append (m1)
        
        b1 = gfg.SubElement(m1, "data")
        b1.text = (cell_obj.value)
                

        searchThis = cell_obj.value
        teacherName = searchThis
        searchThis = searchThis.replace(" ", "")

        

        if searchThis in dir_list:
          #skip, file already exists
          a = 1
          
        else: 
          #create the html file 
          f = open((searchThis), "w")
          #Write information into file here

          
          added.append(cell_obj.value)
          
          f.close()
          

          
         

  i = i+1
  





tree = gfg.ElementTree(root)
#deletes old xml file
os.chdir('../')
os.remove('database.xml')
with open ("database.xml", "wb") as files :
  tree.write(files)



tree = ElementTree.parse("database.xml")   
olp = tree.findall(".//data")
fileList = [t.text for t in olp]#fileList is a list of the pages in the database

#dir_list is a list of pages 
#email-- take whatever item is being deleted from dir_list and match it with the one in fileList. For example, if Scott Bakkum was being removed you could email fileList[1] because it's the same as dir_list, just with spaces
fileListt = []
i = 0
while (i<len(fileList)):#strip out spaces from list of files 
  fileListt.append(fileList[i].replace(' ', ''))
  i = i+1

fileList = fileListt

#Now we loop through the dir_list, any files that are in dir_list and not 'fileList' (the database) is unused and unnescary
i = 0
remove_list = []

while (i<len(dir_list)):#looping through dir_list and asking if each item is in fileList, if it is, nothing happens. If it's not, it's added to remove_list
  if dir_list[i] in (fileList):
    a = 1
  else:
    remove_list.append(dir_list[i])
  i=i+1

i = 0
os.chdir("Data")
while i<len(remove_list):#deletes all files in remove_list
  os.remove(remove_list[i])
  i = i+1

import update